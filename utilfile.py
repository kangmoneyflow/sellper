import os
from openpyxl import Workbook, load_workbook

class ExcelHandler:
    def __init__(self, filepath):
        """
        파일 경로를 받아 기존 엑셀 파일을 로드하거나, 없으면 해당 경로에 새로 생성합니다.
        """
        self.filepath = filepath
        self.wb = None
        self.ws = None

        if os.path.exists(self.filepath):
            self.load_workbook()
        else:
            self.create_new_file()

    def create_new_file(self):
        """
        새로운 엑셀 파일을 사용자가 지정한 경로에 생성합니다.
        """
        print(f"파일이 존재하지 않습니다. 다음 경로에 파일을 생성합니다: {self.filepath}")
        self.create_workbook("Sheet1")
        self.save()

    def create_workbook(self, sheet_name):
        """
        새로운 엑셀 워크북과 시트를 생성합니다.
        """
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_name
        print(f"Created workbook with sheet: {sheet_name}")

    def load_workbook(self):
        """
        기존 엑셀 파일을 로드합니다.
        """
        self.wb = load_workbook(self.filepath)
        self.ws = self.wb.active
        print(f"Loaded existing workbook: {self.filepath}")

    def add_headers(self, headers):
        """
        엑셀 시트에 헤더를 추가합니다.
        """
        if self.ws:
            self.ws.append(headers)
            print(f"Added headers: {headers}")
        else:
            print("Workbook or sheet is not initialized.")

    def append_row(self, row):
        """
        엑셀 시트에 데이터를 한 행씩 추가합니다.
        """
        if self.ws:
            self.ws.append(row)
            print(f"Appended row: {row}")
        else:
            print("Workbook or sheet is not initialized.")

    def save(self):
        """
        엑셀 파일을 저장합니다.
        """
        if self.wb:
            self.wb.save(self.filepath)
            print(f"Saved Excel file: {self.filepath}")
        else:
            print("Workbook is not initialized.")

    def close(self):
        """
        엑셀 파일을 닫습니다.
        """
        if self.wb:
            self.wb.close()
            print(f"Closed Excel file: {self.filepath}")
        else:
            print("Workbook is not initialized.")
